import io
import csv
from typing import Optional, Literal
from fastapi import APIRouter, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import StreamingResponse
from datetime import datetime

from app.core.database import get_db
from app.core.security import get_current_user_id
from app.services.transaction_service import TransactionService
from app.models.transaction import TransactionType, TransactionSource

router = APIRouter()


def _write_csv_data(writer, transactions):
    """Write transaction data to CSV/TSV writer."""
    for t in transactions:
        amount = float(t.amount) if t.amount else 0
        if t.type == TransactionType.EXPENSE:
            amount = -amount
        writer.writerow([
            amount, t.category_name or "", t.description or "",
            t.transaction_date.isoformat() if t.transaction_date else "",
            t.created_at.isoformat() if t.created_at else ""
        ])


def _write_grouped_data_with_totals(writer, grouped_data):
    """Write grouped data to CSV/TSV writer with totals row."""
    total_income = 0.0
    total_expense = 0.0
    
    for item in grouped_data:
        if item['income'] is not None:
            total_income += item['income']
        if item['expense'] is not None:
            total_expense += item['expense']
        writer.writerow([
            item['income'] if item['income'] is not None else "",
            item['expense'] if item['expense'] is not None else "",
            item['category'],
            item['descriptions']
        ])
    
    # Write totals row
    writer.writerow(["", "", "", ""])  # Empty row
    writer.writerow([total_income, total_expense, "TOTAL", ""])


@router.get("/csv")
async def export_csv(
    format: Literal["csv", "tsv", "xlsx"] = Query("csv"),
    type: Optional[TransactionType] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    grouped: Optional[str] = Query(None),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """Export transactions to CSV, TSV, or XLSX format."""
    is_grouped = grouped == 'true'
    
    if is_grouped:
        grouped_data = await TransactionService.get_grouped_for_export(
            db, current_user_id, type, date_from, date_to
        )
        
        if format == "tsv":
            output = io.StringIO()
            writer = csv.writer(output, delimiter='\t', lineterminator='\n')
            writer.writerow(["Income", "Expense", "Category", "Descriptions"])
            _write_grouped_data_with_totals(writer, grouped_data)
            output.seek(0)
            filename = f"transactions_grouped_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.tsv"
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                media_type="text/tab-separated-values",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        if format == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Grouped"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col, header in enumerate(["Income", "Expense", "Category", "Descriptions"], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            total_income = 0.0
            total_expense = 0.0
            for row_idx, item in enumerate(grouped_data, 2):
                income_val = item['income'] if item['income'] is not None else None
                if income_val:
                    total_income += income_val
                cell = ws.cell(row=row_idx, column=1, value=income_val)
                cell.border = thin_border
                if income_val:
                    cell.font = Font(color="008000")
                expense_val = item['expense'] if item['expense'] is not None else None
                if expense_val:
                    total_expense += expense_val
                cell = ws.cell(row=row_idx, column=2, value=expense_val)
                cell.border = thin_border
                if expense_val:
                    cell.font = Font(color="C00000")
                ws.cell(row=row_idx, column=3, value=item['category']).border = thin_border
                desc_cell = ws.cell(row=row_idx, column=4, value=item['descriptions'])
                desc_cell.border = thin_border
                desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Add totals row
            total_row = len(grouped_data) + 3
            ws.cell(row=total_row, column=1, value="").border = thin_border
            ws.cell(row=total_row, column=2, value="").border = thin_border
            ws.cell(row=total_row, column=3, value="").border = thin_border
            ws.cell(row=total_row, column=4, value="").border = thin_border
            
            total_row += 1
            total_income_cell = ws.cell(row=total_row, column=1, value=total_income if total_income else None)
            total_income_cell.border = thin_border
            total_income_cell.font = Font(bold=True, color="008000")
            total_expense_cell = ws.cell(row=total_row, column=2, value=total_expense if total_expense else None)
            total_expense_cell.border = thin_border
            total_expense_cell.font = Font(bold=True, color="C00000")
            total_label_cell = ws.cell(row=total_row, column=3, value="TOTAL")
            total_label_cell.border = thin_border
            total_label_cell.font = Font(bold=True)
            ws.cell(row=total_row, column=4, value="").border = thin_border
            
            ws.freeze_panes = "A2"
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"transactions_grouped_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
        
        # CSV grouped
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Income", "Expense", "Category", "Descriptions"])
        _write_grouped_data_with_totals(writer, grouped_data)
        output.seek(0)
        filename = f"transactions_grouped_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8-sig')), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})
    
    # Regular export
    transactions = await TransactionService.get_all_for_export(db, current_user_id, type, date_from, date_to)
    headers = ["Amount", "Category", "Description", "Transaction Date", "Created At"]
    
    if format == "tsv":
        # TSV format for easy copy-paste into Google Sheets
        output = io.StringIO()
        writer = csv.writer(output, delimiter='\t', lineterminator='\n')
        writer.writerow(headers)
        _write_csv_data(writer, transactions)
        
        output.seek(0)
        filename = f"transactions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.tsv"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/tab-separated-values",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "xlsx":
        # Excel format
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, t in enumerate(transactions, 2):
            # Amount: positive for income, negative for expense
            amount = float(t.amount) if t.amount else 0
            if t.type == TransactionType.EXPENSE:
                amount = -amount
            
            row_data = [
                amount,
                t.category_name or "",
                t.description or "",
                t.transaction_date.isoformat() if t.transaction_date else "",
                t.created_at.isoformat() if t.created_at else ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color negative amounts in red
                if col_idx == 1 and isinstance(value, (int, float)) and value < 0:
                    cell.font = Font(color="C00000")
                
                # Align amount to right
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"transactions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        # Default CSV format
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        _write_csv_data(writer, transactions)
        
        output.seek(0)
        filename = f"transactions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
